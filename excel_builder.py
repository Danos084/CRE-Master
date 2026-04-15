"""
Build / refresh the Excel workbook from listings.json
Two sheets:
  1. All Listings  – full dataset, sortable
  2. New This Week – listings first_seen in last 7 days
"""

import json
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

DATA_DIR = Path("data")
LISTINGS_FILE = DATA_DIR / "listings.json"
EXCEL_OUT = DATA_DIR / "Bayside_Commercial_Listings.xlsx"

# ── Precision Property brand colours ─────────────────────────────────────────
NAVY   = "0D2137"   # header bg
GOLD   = "C9A84C"   # accent
WHITE  = "FFFFFF"
LIGHT  = "F4F6FA"   # alternating row
MID    = "D9DDE6"   # border

COLS = [
    ("Address",        40),
    ("Suburb",         18),
    ("Type",           20),
    ("Size",           14),
    ("Asking Rental",  22),
    ("Listing Agent",  28),
    ("Source",         30),
    ("Link",           50),
    ("First Seen",     14),
    ("Last Updated",   14),
]

FIELDS = [
    "address", "suburb", "type", "size",
    "asking_rental", "listing_agent", "source",
    "link", "first_seen", "last_updated",
]


def _header_style(cell, col_idx):
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=GOLD)
    cell.border = Border(bottom=thin)


def _write_sheet(ws, rows: list[dict], title: str):
    ws.title = title
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Headers
    for ci, (col_name, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        _header_style(cell, ci)
        ws.column_dimensions[get_column_letter(ci)].width = width

    thin = Side(style="thin", color=MID)
    border = Border(bottom=Side(style="thin", color=MID))

    for ri, listing in enumerate(rows, start=2):
        fill_color = LIGHT if ri % 2 == 0 else WHITE
        fill = PatternFill("solid", fgColor=fill_color)

        for ci, field in enumerate(FIELDS, start=1):
            val = listing.get(field, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 1))
            cell.border = border

        # Make link clickable
        link_cell = ws.cell(row=ri, column=8)
        if link_cell.value and link_cell.value.startswith("http"):
            link_cell.hyperlink = link_cell.value
            link_cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")
            link_cell.value = "View Listing"

        ws.row_dimensions[ri].height = 18

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def build_excel(listings: dict) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    all_rows = sorted(listings.values(), key=lambda x: x.get("suburb", ""))
    today = datetime.today()
    cutoff = (today - timedelta(days=7)).strftime("%Y-%m-%d")
    new_rows = [r for r in all_rows if r.get("first_seen", "") >= cutoff]

    ws_all = wb.create_sheet("All Listings")
    _write_sheet(ws_all, all_rows, "All Listings")

    ws_new = wb.create_sheet("New This Week")
    _write_sheet(ws_new, new_rows, "New This Week")

    # Summary sheet
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    summary_rows = [
        ("Report Generated", today.strftime("%d %b %Y %H:%M")),
        ("Total Listings", len(all_rows)),
        ("New This Week", len(new_rows)),
        ("Suburbs Covered", len({r.get("suburb") for r in all_rows})),
        ("Sources", "realcommercial.com.au, commercialrealestate.com.au"),
    ]
    for ri, (label, val) in enumerate(summary_rows, start=2):
        ws_sum.cell(row=ri, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
        ws_sum.cell(row=ri, column=2, value=val).font = Font(name="Arial", size=11)

    ws_sum["A1"] = "Bayside Commercial Listings"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=18, color=NAVY)
    ws_sum["A1"].fill = PatternFill("solid", fgColor=LIGHT)
    ws_sum.row_dimensions[1].height = 36

    wb.save(EXCEL_OUT)
    print(f"Excel saved → {EXCEL_OUT}  ({len(all_rows)} listings)")
    return EXCEL_OUT


if __name__ == "__main__":
    with open(LISTINGS_FILE) as f:
        data = json.load(f)
    build_excel(data)
